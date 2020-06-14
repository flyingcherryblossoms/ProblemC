import random
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

# # 读取Excel
# sourceFile = pd.ExcelFile("../题目/附件1：仓库数据.xlsx")
# CargoContainer = pd.read_excel(sourceFile, "货格")
# taskLists = pd.read_excel(sourceFile, "任务单")  # 任务单
# ReviewStation = pd.read_excel(sourceFile, "复核台")  # 复核台
# distanceP2 = pd.ExcelFile("./P2_distance.xlsx")
#
# # 任务单T0001
# taskList = []
# for i in range(23):
#     taskList.append(list(taskLists.iloc[i]))
#     taskList[i].append((int(taskList[i][2][-5:-2]) - 1) * 15 + int(taskList[i][2][-2:]))
# for i in range(1, 14):
#     if i < 10:
#         taskList.append(['T0001', '', 'FH0' + str(i), 0, 3000 + i])
#     else:
#         taskList.append(['T0001', '', 'FH' + str(i), 0, 3000 + i])
# # for i in range(len(taskList)):
# #     print(i + 1, taskList[i])
#
# # 任务点编号
# taskName = []
# for i in range(len(taskList)):
#     taskName.append(taskList[i][2])
# # print(len(taskName), taskName)
#
# # 货格和复核台的位置
# taskPoints = []
# for i in range(23):
#     for j in range(len(CargoContainer)):
#         if taskList[i][2] == list(CargoContainer.iloc[j])[0]:
#             taskPoints.append([int(list(CargoContainer.iloc[j])[1]), int(list(CargoContainer.iloc[j])[2])])
# for i in range(len(ReviewStation)):
#     taskPoints.append([int(list(ReviewStation.iloc[i])[1]), int(list(ReviewStation.iloc[i])[2])])
# taskPoints = np.array(taskPoints)
#
# # 任务单每个货格到其它任务单货格的距离和所有复核台的距离
# DistanceAll = []
# distanceP2Data = pd.read_excel(distanceP2)
# for i in range(len(distanceP2Data)):
#     DistanceAll.append(list(distanceP2Data.iloc[i]))
#     # print(len(Distance[i]), Distance[i])

def readExcel():
    # 读取Excel
    sourceFile = pd.ExcelFile("../题目/附件1：仓库数据.xlsx")
    CargoContainer = pd.read_excel(sourceFile, "货格")
    taskLists = pd.read_excel(sourceFile, "任务单")  # 任务单
    ReviewStation = pd.read_excel(sourceFile, "复核台")  # 复核台
    distanceP2 = pd.ExcelFile("./P2_distance.xlsx")

    # 任务单T0001
    taskList = []
    for i in range(23):
        taskList.append(list(taskLists.iloc[i]))
        taskList[i].append((int(taskList[i][2][-5:-2]) - 1) * 15 + int(taskList[i][2][-2:]))
    for i in range(1, 14):
        if i < 10:
            taskList.append(['T0001', '', 'FH0' + str(i), 0, 3000 + i])
        else:
            taskList.append(['T0001', '', 'FH' + str(i), 0, 3000 + i])

    # 货格和复核台的位置
    taskPoint = []
    for i in range(23):
        for j in range(len(CargoContainer)):
            if taskList[i][2] == list(CargoContainer.iloc[j])[0]:
                taskPoint.append([int(list(CargoContainer.iloc[j])[1]), int(list(CargoContainer.iloc[j])[2])])
    for i in range(len(ReviewStation)):
        taskPoint.append([int(list(ReviewStation.iloc[i])[1]), int(list(ReviewStation.iloc[i])[2])])
    taskPoint = np.array(taskPoint)

    # 任务单每个货格到其它任务单货格的距离和所有复核台的距离
    Distance = []
    distanceP2Data = pd.read_excel(distanceP2)
    for i in range(len(distanceP2Data)):
        Distance.append(list(distanceP2Data.iloc[i]))
        # print(len(Distance[i]), Distance[i])
    return taskList, taskPoint, Distance


taskList, taskPoints, DistanceAll = readExcel()

# 种群数
count = 100
# 改良次数
improve_count = 1000
# 进化次数
itter_time = 500
# 设置强者的定义概率，即种群前30%为强者
retain_rate = 0.3
# 设置弱者的存活概率
random_select_rate = 0.5
# 变异率
mutation_rate = 0.1
# 设置起点
origin = 32  # 起点FH10
index = [i for i in range(24)]
index.remove(23)


# 总距离
def get_total_distance(x):
    distance = 0
    distance += DistanceAll[origin][x[0]]
    for i in range(len(x)):
        if i == len(x) - 1:
            distance += DistanceAll[origin][x[i]]
        else:
            distance += DistanceAll[x[i]][x[i + 1]]
    return distance


# 改良圈子算法
def improve(x):
    indexi = 0
    distance = get_total_distance(x)
    while indexi < improve_count:
        # randint [a,b]
        u = random.randint(0, len(x) - 1)
        v = random.randint(0, len(x) - 1)
        if u != v:
            new_x = x.copy()
            t = new_x[u]
            new_x[u] = new_x[v]
            new_x[v] = t
            new_distance = get_total_distance(new_x)
            if new_distance < distance:
                distance = new_distance
                x = new_x.copy()
        else:
            continue
        indexi += 1
    return x


# 自然选择
def selection(population):
    """
    选择
    先对适应度从大到小排序，选出存活的染色体
    再进行随机选择，选出适应度虽然小，但是幸存下来的个体
    """
    # 对总距离从小到大进行排序
    graded = [[get_total_distance(x), x] for x in population]
    graded = [x[1] for x in sorted(graded)]
    # 选出适应性强的染色体
    retain_length = int(len(graded) * retain_rate)
    parents = graded[:retain_length]
    # 选出适应性不强，但是幸存的染色体
    for chromosome in graded[retain_length:]:
        if random.random() < random_select_rate:
            parents.append(chromosome)
    return parents


# 交叉繁殖
def crossover(parents):
    # 生成子代的个数,以此保证种群稳定
    target_count = count - len(parents)
    # 孩子列表
    children = []
    while len(children) < target_count:
        male_index = random.randint(0, len(parents) - 1)
        female_index = random.randint(0, len(parents) - 1)
        if male_index != female_index:
            male = parents[male_index]
            female = parents[female_index]

            left = random.randint(0, len(male) - 2)
            right = random.randint(left + 1, len(male) - 1)

            # 交叉片段
            gene1 = male[left:right]
            gene2 = female[left:right]

            child1_c = male[right:] + male[:right]
            child2_c = female[right:] + female[:right]
            child1 = child1_c.copy()
            child2 = child2_c.copy()

            for o in gene2:
                child1_c.remove(o)

            for o in gene1:
                child2_c.remove(o)

            child1[left:right] = gene2
            child2[left:right] = gene1

            child1[right:] = child1_c[0:len(child1) - right]
            child1[:left] = child1_c[len(child1) - right:]

            child2[right:] = child2_c[0:len(child1) - right]
            child2[:left] = child2_c[len(child1) - right:]

            children.append(child1)
            children.append(child2)

    return children


# 变异
def mutation(children):
    for i in range(len(children)):
        if random.random() < mutation_rate:
            child = children[i]
            u = random.randint(1, len(child) - 4)
            v = random.randint(u + 1, len(child) - 3)
            w = random.randint(v + 1, len(child) - 2)
            child = children[i]
            child = child[0:u] + child[v:w] + child[u:v] + child[w:]
    return children


# 得到最佳纯输出结果
def get_result(population):
    graded = [[get_total_distance(x), x] for x in population]
    graded = sorted(graded)
    return graded[0][0], graded[0][1]


# 总距离
def get_distance(x):
    distance = 0
    for k in range(len(x) - 1):
        distance += DistanceAll[x[k]][x[k + 1]]
    return distance


def draw(Path, minRegister):
    X = []
    Y = []
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.rcParams.update({'font.size': 8})
    for index in Path:
        x = taskPoints[index, 0]
        y = taskPoints[index, 1]
        X.append(x)
        Y.append(y)
        plt.annotate(taskList[index][2], xy=(x, y), xytext=(x + 600, y))
    plt.plot(X, Y, '-o')
    plt.show()
    plt.plot(list(range(len(minRegister))), minRegister)
    plt.show()


minPath = []  # 最短路径
minDistance = 9999999999  # 最短路径长度
minRegister = []  # 最短的那次的迭代过程的路径变化
# 多次计算，取最短的一次
for times in range(5):
    # 使用改良圈算法初始化种群
    population = []
    for i in range(count):
        # 随机生成个体
        x = index.copy()
        random.shuffle(x)
        x = improve(x)
        population.append(x)

    register = []  # 记录遗传过程的每一组序列的总长度变化
    i = 0
    distance, result_path = get_result(population)
    while i < itter_time:
        # 选择繁殖个体群
        parents = selection(population)
        # 交叉繁殖
        children = crossover(parents)
        # 变异操作
        children = mutation(children)
        # 更新种群
        population = parents + children

        distance, result_path = get_result(population)
        register.append(distance)
        i = i + 1

    print("当前路径：", distance, result_path)

    result_path = [origin] + result_path + [origin]
    draw(result_path, register)

    if minDistance > distance:
        minDistance = distance
        minPath = result_path
        minRegister = register

    tempPath = minPath.copy()  # 算出的最小回路
    tempPathReverse = minPath.copy()  # 首尾倒置的最小回路
    tempPathReverse.reverse()
    for i in range(13):  # 去掉最后一个点（终点，或者起点，因为是同一个点，所以要倒过来算一次），环取最短路径，一共13个复核台，26种情况
        ccPath = []  # 记录货格名称
        tempPath.pop()  # 去尾
        tempPath.append(i + 23)  # 加复核台
        tempPathReverse.pop()  # 去尾
        tempPathReverse.append(i + 23)  # 加复核台
        pathlenth1 = get_distance(tempPath)
        pathlenth2 = get_distance(tempPathReverse)
        print(pathlenth1, pathlenth2, 'FH', i + 1)
        if pathlenth1 < pathlenth2:
            if minDistance > pathlenth1:
                minDistance = pathlenth1
                minPath = tempPath.copy()
                print("当前minDistance", minDistance, "当前minPath：", minPath)
        else:
            if minDistance > pathlenth2:
                minDistance = pathlenth2
                minPath = tempPathReverse.copy()
                print("当前minDistance", minDistance, "当前minPath：", minPath)
        # print(tempPath)
        # print(tempPathReverse)
        # for j in range(len(minPath)):
        #     ccPath.append(taskList[minPath[j]][2])
        # print(ccPath)  # 输出最短路径的货格代码


draw(minPath, minRegister)
print("\n最短距离：", minDistance)
print("最短路径：", len(minPath), minPath)
ccPath = []
outputData = [["序号", "元素原始名称", "元素序号(1-3013)", "商品件数"]]  # 输出到表格的内容
for i in range(len(minPath)):
    ccPath.append(taskList[minPath[i]][2])
    outputData.append([i + 1, taskList[minPath[i]][2], taskList[minPath[i]][4], taskList[minPath[i]][3]])
print(ccPath)
# for i in range(len(outputData)):
#     print(outputData[i])


# 写入Excel
writer = pd.ExcelWriter("Ques2.xlsx", engine='openpyxl')
book = openpyxl.load_workbook(writer.path)
writer.book = book
# 清除原来的数据
idx = book.sheetnames.index('Ques2')
book.remove(book.worksheets[idx])
# book.create_sheet('Ques2', idx)
# 写入新的数据
pd.DataFrame(outputData).to_excel(excel_writer=writer, sheet_name='Ques2', index=None, header=False)
writer.save()
writer.close()


# 计算花费的时间
# 时间包括复核台复核打包时间，路程上的时间，取货的时间
Total_Time = float(minDistance)/1500  #初始值为走完全程的时间
for i in range(len(taskList)):
    amount = taskList[i][3]
    time = 0
    if amount >= 3:
        time = 4 * amount
    else:
        time = 5 * amount
    Total_Time += time
Total_Time += 30    # 复核台复核打包时间
print("花费总时间：", Total_Time, "s")

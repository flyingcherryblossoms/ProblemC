import openpyxl
import pandas as pd
import numpy as np
from GATSP import GATSP


def writeToExcel(fileName, outputData):
    # 写入Excel
    writer = pd.ExcelWriter(fileName+'.xlsx', engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    # 清除原来的数据
    try:
        idx = book.sheetnames.index(fileName)
        book.remove(book.worksheets[idx])
        # book.create_sheet('Ques2', idx)
        # 写入新的数据
    except:
        pass
    pd.DataFrame(outputData).to_excel(excel_writer=writer, sheet_name=fileName, index=None, header=False)
    writer.save()
    writer.close()


def cal_time(minDistance, taskList):
    # 计算花费的时间
    # 时间包括复核台复核打包时间，路程上的时间，取货的时间
    Total_Time = float(minDistance) / 1500  # 初始值为走完全程的时间
    for i in range(len(taskList)):
        amount = taskList[i][3]
        if amount >= 3:
            time = 4 * amount
        else:
            time = 5 * amount
        Total_Time += time
    Total_Time += 30  # 复核台复核打包时间
    print("花费总时间：", Total_Time, "s")
    return Total_Time


def readExcel():
    # 读取Excel
    sourceFile = pd.ExcelFile("../题目/附件1：仓库数据.xlsx")
    CargoContainer = pd.read_excel(sourceFile, "货格")
    taskLists = pd.read_excel(sourceFile, "任务单")  # 任务单
    ReviewStation = pd.read_excel(sourceFile, "复核台")  # 复核台
    distanceP2 = pd.ExcelFile("./P2_distance.xlsx")

    # 任务单T0001
    taskList = []
    for i in range(23):
        taskList.append(list(taskLists.iloc[i]))
        taskList[i].append((int(taskList[i][2][-5:-2]) - 1) * 15 + int(taskList[i][2][-2:]))
    for i in range(1, 14):
        if i < 10:
            taskList.append(['T0001', '', 'FH0' + str(i), 0, 3000 + i])
        else:
            taskList.append(['T0001', '', 'FH' + str(i), 0, 3000 + i])

    # 货格和复核台的位置
    taskPoint = []
    for i in range(23):
        for j in range(len(CargoContainer)):
            if taskList[i][2] == list(CargoContainer.iloc[j])[0]:
                taskPoint.append([int(list(CargoContainer.iloc[j])[1]), int(list(CargoContainer.iloc[j])[2])])
    for i in range(len(ReviewStation)):
        taskPoint.append([int(list(ReviewStation.iloc[i])[1]), int(list(ReviewStation.iloc[i])[2])])
    taskPoint = np.array(taskPoint)

    # 任务单每个货格到其它任务单货格的距离和所有复核台的距离
    Distance = []
    distanceP2Data = pd.read_excel(distanceP2)
    for i in range(len(distanceP2Data)):
        Distance.append(list(distanceP2Data.iloc[i]))
        # print(len(Distance[i]), Distance[i])
    return taskList, taskPoint, Distance


taskList, taskPoint, Distance = readExcel()


# 种群数
pop_count = 100
# 改良次数
improve_count = 1000
# 进化次数
itter_time = 500
# 设置强者的定义概率，即种群前30%为强者
retain_rate = 0.3
# 设置弱者的存活概率
random_select_rate = 0.5
# 变异率
mutation_rate = 0.1
# 设置起点
origin = 32  # 起点FH10
# 点的个数，不含首尾
point_count = 23
# 计算的所有坐标序列，不包含首尾
index = [i for i in range(point_count)]

# 遗传算法
GA = GATSP(point_count=point_count, pop_count=pop_count, improve_count=improve_count, itter_time=itter_time,
           retain_rate=retain_rate, random_select_rate=random_select_rate, mutation_rate=mutation_rate,
           origin=origin, index=index, taskList=taskList, taskPoint=taskPoint, Distance=Distance)
RS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]     # 复核台：0表示第一个，后面依此类推
minDistance, minPath, ccPath, minRegister = GA.run(100, RS)
GA.draw(minPath, minRegister)


print("\n最短距离：", minDistance)
print("最短路径：", len(minPath), minPath)
print(ccPath)


outputData = [["序号", "元素原始名称", "元素序号(1-3013)", "商品件数"]]  # 输出到表格的内容
for i in range(len(minPath)):
    outputData.append([i + 1, taskList[minPath[i]][2], taskList[minPath[i]][4], taskList[minPath[i]][3]])
# for i in range(len(outputData)):
#     print(outputData[i])

writeToExcel("Ques2_", outputData=outputData)

cal_time(minDistance, taskList)




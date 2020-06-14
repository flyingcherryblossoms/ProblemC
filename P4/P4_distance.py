import pandas as pd
import openpyxl
import os


def writeToExcel(fileName, outputData):
    filename = './src/'+fileName+'.xlsx'
    if not os.path.exists(filename):
        pd.DataFrame().to_excel(filename)

    # 写入Excel
    print(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    # 清除原来的数据
    try:
        idx = book.sheetnames.index(fileName)
        book.remove(book.worksheets[idx])
        # book.create_sheet('Ques2', idx)
        # 写入新的数据
    except:
        pass
    pd.DataFrame(outputData).to_excel(excel_writer=writer, sheet_name=fileName, index=None)
    writer.save()
    writer.close()

# 读取Excel
distanceFile = pd.ExcelFile("../P1/Ques1.xlsx")
sourceFile = pd.ExcelFile("../题目/附件1：仓库数据.xlsx")
distanceData = pd.read_excel(distanceFile)  # 距离数据
ReviewStation = pd.read_excel(sourceFile, "复核台")
CargoContainer = pd.read_excel(sourceFile, "货格")
taskListAll = pd.read_excel(sourceFile, "任务单")  # 任务单

taskLists = []  # 任务单
taskName = []
for i in range(1, 50):
    if i < 10:
        taskName.append('T000'+str(i))
    else:
        taskName.append('T00'+str(i))

for i in range(len(taskName)):
    print(i, taskName[i])
    taskList = []
    k = 0
    for j in range(len(taskListAll)):
        if list(taskListAll.iloc[j])[0] == taskName[i]:
            taskList.append(list(taskListAll.iloc[j]))
            # print("货架：", int(taskList[k][2][-5:-2]))
            # print("货格", int(taskList[k][2][-2:]))
            # print("序号：", (int(taskList[k][2][-5:-2])-1)*15 + int(taskList[k][2][-2:]))
            taskList[k].append((int(taskList[k][2][-5:-2]) - 1) * 15 + int(taskList[k][2][-2:]))
            print(taskList[k])
            k += 1
    for j in range(1, 14):
        if j < 10:
            taskList.append(['', '', 'FH0' + str(j), 0, 3000 + j])
        else:
            taskList.append(['', '', 'FH' + str(j), 0, 3000 + j])
    writeToExcel('P4_taskLists'+str(i+1), taskList)
    taskLists.append(taskList)

# 货格和复核台的位置
for k in range(len(taskLists)):
    taskPoint = []
    for i in range(len(taskLists[k])):
        for j in range(len(CargoContainer)):
            if taskLists[k][i][2] == list(CargoContainer.iloc[j])[0]:
                taskPoint.append([int(list(CargoContainer.iloc[j])[1]), int(list(CargoContainer.iloc[j])[2])])
    for i in range(len(ReviewStation)):
        taskPoint.append([int(list(ReviewStation.iloc[i])[1]), int(list(ReviewStation.iloc[i])[2])])
    print(taskPoint)
    writeToExcel("P4_taskPoint"+str(k+1), taskPoint)
    # taskPoint = np.array(taskPoint)

distances = []  # 任务单每个货格到其它任务单货格的距离和所有复核台的距离
for k in range(len(taskLists)):
    distance = []
    # 保存计算所需数据，避免去所有距离中读取，太浪费时间了
    for i in range(len(taskLists[k])):
        distancei = []
        indexi = taskLists[k][i][4] - 1
        # print("indexi= ", indexi)
        for j in range(len(taskLists[k])):
            indexj = taskLists[k][j][4] - 1
            # print("indexj= ", indexj)
            distancei.append(list(distanceData.iloc[indexi])[indexj])
        for j in range(3000, 3013):
            indexj = j
            distancei.append(list(distanceData.iloc[indexi])[indexj])
        print(distancei)
        distance.append(distancei)
    # writeToExcel('P4_distance' + str(k+1), distance)
    distances.append(distance)


def Manhattan(x1, y1, x2, y2):
    # 计算曼哈顿距离
    return abs(int(x1) - int(x2)) + abs(int(y1) - int(y2))

# 计算复核台之间距离
ans2 = []
for i in range(len(ReviewStation)):
    ansi = []
    rsi = list(ReviewStation.iloc[i])  # 复核台i信息
    for j in range(len(ReviewStation)):
        rsj = list(ReviewStation.iloc[j])  # 复核台j信息
        dist = Manhattan(rsi[1], rsi[2], rsj[1], rsj[2])
        ansi.append(dist)
    ans2.append(ansi)
    # print(len(ansi), ansi)

for k in range(len(distances)):
    lenansj = len(distances[k])
    lenansi = len(distances[k][0])
    t = 0
    for i in range(lenansi - 13, lenansi):  # 列循环
        ansi = []
        for j in range(0, lenansj):  # 行循环
            ansi.append(distances[k][j][i])
        ansi.extend(ans2[t])
        t += 1
        # print(len(ansi), ansi)
        distances[k].append(ansi)

for k in range(len(taskLists)):
    writeToExcel("P4_distance" + str(k+1), distances[k])




import os
import openpyxl
import pandas as pd
import numpy as np
from GATSP import GATSP


def writeToExcel(fileName, outputData):
    filename = fileName + '.xlsx'
    if not os.path.exists(filename):
        pd.DataFrame().to_excel(filename)
    # 写入Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    # 清除原来的数据
    try:
        idx = book.sheetnames.index(fileName)
        book.remove(book.worksheets[idx])
        # book.create_sheet('Ques2', idx)
        # 写入新的数据
    except:
        pass
    pd.DataFrame(outputData).to_excel(excel_writer=writer, sheet_name=fileName, index=None, header=False)
    writer.save()
    writer.close()


def cal_time(minDistance, taskList):
    # 计算花费的时间
    # 时间包括复核台复核打包时间，路程上的时间，取货的时间
    Total_Time = float(minDistance) / 1500  # 初始值为走完全程的时间
    for i in range(len(taskList)):
        amount = taskList[i][3]
        if amount >= 3:
            time = 4 * amount
        else:
            time = 5 * amount
        Total_Time += time
    Total_Time += 30  # 复核台复核打包时间
    # print("花费总时间：", Total_Time, "s")
    return Total_Time


def readExcel(filename):
    # 读取Excel
    file = pd.ExcelFile("./src/" + filename + ".xlsx")
    # 任务单T0001
    data = []
    fileReader = pd.read_excel(file, sheet_name=filename)
    for i in range(len(fileReader)):
        data.append(list(fileReader.iloc[i]))
        # print(len(data[i]), data[i])
    return data


taskLists = []
taskPoints = []
distances = []
for indexi in range(1, 50):
    taskLists.append(readExcel("P4_taskLists" + str(indexi)))
    taskPoints.append(readExcel("P4_taskPoint" + str(indexi)))
    distances.append(readExcel("P4_distance" + str(indexi)))
# print(taskLists)
# print(taskPoints)
# print(distances)

# 种群数
pop_count = 100
# 改良次数
improve_count = 2000
# 进化次数
itter_time = 200
# 设置强者的定义概率，即种群前30%为强者
retain_rate = 0.3
# 设置弱者的存活概率
random_select_rate = 0.5
# 变异率
mutation_rate = 0.1

origins = [0, 2, 9, 11]  # 记录每一个点的起点，初始起点为FH03
minDistances = []  # 记录每个货单的最短路径长度
minPaths = []  # 记录每个货单的最短路径
minReisters = []  # 记录每个货单最短路径的迭代过程
ccPaths = []  # 记录每个货单最短路径上的货格编号
for indexi in range(len(taskLists)):
    # 点的个数，不含首尾
    point_count = len(taskLists[indexi]) - 13
    # print("点个数", point_count)
    # 计算的所有坐标序列，不包含首尾
    index = [i for i in range(point_count)]
    minDistance = 99999999
    minPath = []
    ccPath = []
    minRegister = []
    index2 = 0
    # 分别从四个复核台出发，算最短路径
    for indexj in range(len(origins)):
        origin = origins[indexj] + point_count
        # 遗传算法
        GA = GATSP(point_count=point_count, pop_count=pop_count, improve_count=improve_count, itter_time=itter_time,
                   retain_rate=retain_rate, random_select_rate=random_select_rate, mutation_rate=mutation_rate,
                   origin=origin, index=index, taskList=taskLists[indexi],
                   taskPoint=np.array(taskPoints[indexi]),
                   Distance=distances[indexi])
        RS = [0, 2, 9, 11]  # 复核台：0表示第一个，后面依此类推
        minDistancej, minPathj, ccPathj, minRegisterj = GA.run(5, RS)
        # print(minDistancej)
        if minDistance > minDistancej:
            minDistance = minDistancej
            minPath = minPathj
            minRegister = minRegisterj
            ccPath = ccPathj
            index2 = indexj

    GA = GATSP(taskList=taskLists[indexi],
               taskPoint=np.array(taskPoints[indexi]),
               Distance=distances[indexi])
    # GA.draw(minPath, None)

    minDistances.append(minDistance)
    minPaths.append(minPath)
    ccPaths.append(ccPath)
    minReisters.append(minRegister)
    # print(origins)
    print(indexi+1)

# for indexi in range(len(minDistances)):
#     print("\n最短距离：", minDistances[indexi])
#     print("最短路径：", len(minPaths[indexi]), minPaths[indexi])
#     print(ccPaths[indexi])

# 计算每个任务单的时间
minTimes = []
for indexi in range(len(taskLists)):
    minTimes.append(cal_time(minDistances[indexi], taskLists[indexi]))
    # print(minTimes[indexi])

# 按任务单时间排序
for indexi in range(len(minTimes)):
    minIndex = indexi
    for indexj in range(indexi+1, len(minTimes)):
        if minTimes[minIndex] > minTimes[indexj]:
            minIndex = indexj
    if minIndex != indexi:
        tempTime = minTimes[minIndex]
        tempDistance = minDistances[minIndex]
        tempPath = minPaths[minIndex]
        tempccPath = ccPaths[minIndex]
        tempTaskList = taskLists[minIndex]

        minTimes.pop(minIndex)
        minDistances.pop(minIndex)
        minPaths.pop(minIndex)
        ccPaths.pop(minIndex)
        taskLists.pop(minIndex)

        minTimes.insert(indexi, tempTime)
        minDistances.insert(indexi, tempDistance)
        minPaths.insert(indexi, tempPath)
        ccPaths.insert(indexi, tempccPath)
        taskLists.insert(indexi, tempTaskList)


PTime = []  # 工人的工作时间
RSTime = []  # 复核台的时刻，即上一个员工到达的时间
for indexi in range(9):
    PTime.append([[], [0]])  # 第一个[]存储货单编号，第二个为该货单的时间，及货单时间加上等待时间
for indexi in range(4):
    RSTime.append(-30)
RS = [0, 2, 9, 11]  # 货单编号

# 把剩余货单时间最少的货单分配给目前时间最少的工人
for indexi in range(len(taskLists)):
    minPtime = 99999999
    minIndex = 0
    # 寻找用时最少的工人
    for indexj in range(len(PTime)):
        if minPtime > sum(PTime[indexj][1]):
            minIndex = indexj
            minPtime = sum(PTime[indexj][1])
    # 给工人加上工作时间，记录货单号
    PTime[minIndex][1].append(minTimes[indexi])
    PTime[minIndex][0].append(taskLists[indexi][1][0])
    # 判断终点是哪个货台，就在那个货台加上该的时间，并判断工人是否需要等待
    for indexj in range(len(RS)):
        if minPaths[indexi][-1]-len(minPaths[indexi])+2 == RS[indexj]:
            waitTime = 0    # 等待时间
            print("P"+str(minIndex+1)+"到达时间：", sum(PTime[minIndex][1]))
            print("当前任务单：", PTime[minIndex][0][-1])
            print(str(RS[indexj]) + "上次任务单完成时刻：", RSTime[indexj]+30)
            if sum(PTime[minIndex][1]) < RSTime[indexj]+30:
                waitTime = RSTime[indexj] + 30 - sum(PTime[minIndex][1])
                print(PTime[minIndex][0][-1] + "需要等待：", waitTime)
            else:
                print(PTime[minIndex][0][-1] + "不需要等待")
            RSTime[indexj] = sum(PTime[minIndex][1]) + waitTime
            print(RSTime[indexj], end='\n\n')
            break

# 计算最长的员工时间就是出库的时间
outTime = 0
for indexi in range(len(PTime)):
    if outTime < sum(PTime[indexi][1]):
        outTime = sum(PTime[indexi][1])
print("出库时间：", outTime, 's')

# for indexi in range(len(minDistances)):
#     print("\n花费时间：", minTimes[indexi])
#     print("最短距离：", minDistances[indexi])
#     print("最短路径：", len(minPaths[indexi]), minPaths[indexi])
#     print(ccPaths[indexi])

outputData = [["序号", "人(P1-P9)", "任务单（T0001-T0049）", "元素原始名称", "元素序号(1-3013)", "商品件数"]]  # 输出到表格的内容
amount = 1
for indexi in range(len(taskLists)):
    for indexj in minPaths[indexi]:
        for indexk in range(len(PTime)):
            if taskLists[indexi][indexj][0] in PTime[indexk][0] or indexj == minPaths[indexi][0] or indexj == minPaths[indexi][-1]:
                outputData.append([amount, "P"+str(indexk+1),
                                  taskLists[indexi][indexj][0],
                                  taskLists[indexi][indexj][2],
                                  taskLists[indexi][indexj][4],
                                  taskLists[indexi][indexj][3]])
                amount += 1
                break
# for indexi in range(len(outputData)):
#     print(outputData[indexi])
# for i in range(len(outputData)):
#     print(outputData[i])
writeToExcel("Ques4", outputData=outputData)


# 计算复核台利用率
RS = [0, 2, 9, 11]
RSName = ["FH01", "FH03", "FH10", "FH12"]
Timei = []
for indexi in range(len(RS)):
    Timei.append([])
    for indexj in range(len(minPaths)):
        if minPaths[indexj][-1]-len(minPaths[indexj])+2 == RS[indexi]:
            Timei[indexi].append(minTimes[indexj])
for indexi in range(len(Timei)):
    if len(Timei[indexi]) != 0:
        print(str(RSName[indexi]) + "利用率：", float(len(Timei[indexi])*30)/RSTime[indexi])
    else:
        print(str(RSName[indexi]) + "未使用")

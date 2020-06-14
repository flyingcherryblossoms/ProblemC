import os
import openpyxl
import pandas as pd
import numpy as np
from GATSP import GATSP


def writeToExcel(fileName, outputData):
    filename = fileName + '.xlsx'
    if not os.path.exists(filename):
        pd.DataFrame().to_excel(filename)
    # 写入Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    # 清除原来的数据
    try:
        idx = book.sheetnames.index(fileName)
        book.remove(book.worksheets[idx])
        # book.create_sheet('Ques2', idx)
        # 写入新的数据
    except:
        pass
    pd.DataFrame(outputData).to_excel(excel_writer=writer, sheet_name=fileName, index=None, header=False)
    writer.save()
    writer.close()


def cal_time(minDistance, taskList):
    # 计算花费的时间
    # 时间包括复核台复核打包时间，路程上的时间，取货的时间
    Total_Time = float(minDistance) / 1500  # 初始值为走完全程的时间
    for i in range(len(taskList)):
        amount = taskList[i][3]
        if amount >= 3:
            time = 4 * amount
        else:
            time = 5 * amount
        Total_Time += time
    Total_Time += 30  # 复核台复核打包时间
    print("花费总时间：", Total_Time, "s")
    return Total_Time


def readExcel(filename):
    # 读取Excel
    file = pd.ExcelFile("./src/" + filename + ".xlsx")
    # 任务单T0001
    data = []
    fileReader = pd.read_excel(file, sheet_name=filename)
    for i in range(len(fileReader)):
        data.append(list(fileReader.iloc[i]))
        # print(len(data[i]), data[i])
    return data


taskLists = []
taskPoints = []
distances = []
for indexi in range(1, 6):
    taskLists.append(readExcel("P3_taskLists" + str(indexi)))
    taskPoints.append(readExcel("P3_taskPoint" + str(indexi)))
    distances.append(readExcel("P3_distance" + str(indexi)))
# for i in range(len(taskLists)):
#     print(taskLists[i])
#     print(taskPoints[i])
#     print(distances[i])

# 种群数
pop_count = 100
# 改良次数
improve_count = 1000
# 进化次数
itter_time = 500
# 设置强者的定义概率，即种群前30%为强者
retain_rate = 0.3
# 设置弱者的存活概率
random_select_rate = 0.5
# 变异率
mutation_rate = 0.1

origins = [2]  # 记录每一个点的起点，初始起点为FH03
minDistances = []  # 记录每个货单的最短路径长度
minPaths = []  # 记录每个货单的最短路径
minReisters = []  # 记录每个货单最短路径的迭代过程
ccPaths = []  # 记录每个货单最短路径上的货格编号

taskLists2 = taskLists.copy()
taskPoints2 = taskPoints.copy()
distances2 = distances.copy()

taskLists3 = []
taskPoints3 = []
distances3 = []

for indexi in range(len(taskLists)):
    minDistance = 99999999
    minPath = []
    ccPath = []
    minRegister = []
    index2 = 0  # 记住上上一个最短的货单序号，用于删除已经被计算过的单号
    point_count2 = 0  # 记住最短货单的点的个数
    # 找出剩余货单中路径最短的，并且从上一个任务单结束的货台出发
    # print("taskLists2：", len(taskLists2))
    for indexj in range(len(taskLists2)):
        # 点的个数，不含首尾
        point_count = len(taskLists2[indexj]) - 13
        # print("点个数", point_count)
        origin = origins[indexi] + point_count
        # 计算的所有坐标序列，不包含首尾
        index = [i for i in range(point_count)]
        # 遗传算法
        GA = GATSP(point_count=point_count, pop_count=pop_count, improve_count=improve_count, itter_time=itter_time,
                   retain_rate=retain_rate, random_select_rate=random_select_rate, mutation_rate=mutation_rate,
                   origin=origin, index=index, taskList=taskLists2[indexj],
                   taskPoint=np.array(taskPoints2[indexj]),
                   Distance=distances2[indexj])
        RS = [2, 10]  # 复核台：0表示第一个，后面依此类推
        minDistancej, minPathj, ccPathj, minRegisterj = GA.run(10, RS)
        if minDistance > minDistancej:
            minDistance = minDistancej
            minPath = minPathj
            minRegister = minRegisterj
            ccPath = ccPathj
            index2 = indexj
            point_count2 = point_count

    GA = GATSP(taskList=taskLists2[index2],
               taskPoint=np.array(taskPoints2[index2]),
               Distance=distances2[index2])
    GA.draw(minPath, None)
    taskLists3.append(taskLists2[index2])
    taskPoints3.append(taskPoints2[index2])
    distances3.append(distances2[index2])

    taskLists2.pop(index2)
    taskPoints2.pop(index2)
    distances2.pop(index2)

    minDistances.append(minDistance)
    minPaths.append(minPath)
    ccPaths.append(ccPath)
    minReisters.append(minRegister)
    origins.append(minPath[-1] - point_count2)
    # print(origins)

for indexi in range(len(minDistances)):
    print("\n最短距离：", minDistances[indexi])
    print("最短路径：", len(minPaths[indexi]), minPaths[indexi])
    print(ccPaths[indexi])

outputData = [['序号', '任务单(T0002-T0006)', '元素原始名称', '元素序号(1-3013)', '商品件数']]  # 输出到表格的内容
counts = 0
indexi = 0
indexj = 0
for indexi in range(len(minPaths)):
    for indexj in range(len(minPaths[indexi])-1):
        outputData.append([counts+1, taskLists3[indexi][minPaths[indexi][indexj]][0],
                           taskLists3[indexi][minPaths[indexi][indexj]][2],
                         taskLists3[indexi][minPaths[indexi][indexj]][4],
                           taskLists3[indexi][minPaths[indexi][indexj]][3]])
        counts += 1
outputData.append([counts+1, taskLists3[indexi][minPaths[indexi][indexj+1]][0],
                           taskLists3[indexi][minPaths[indexi][indexj+1]][2],
                         taskLists3[indexi][minPaths[indexi][indexj+1]][4],
                           taskLists3[indexi][minPaths[indexi][indexj+1]][3]])
for i in range(len(outputData)):
    print(outputData[i])

writeToExcel("Ques3_", outputData=outputData)

Time = 0
for indexi in range(len(minDistances)):
    Time += cal_time(minDistances[indexi], taskLists3[indexi])
print("出库时间", Time, 's')
